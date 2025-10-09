#!/usr/bin/env python3
"""
CSV Parser for Redmine Time Entries

This script parses the timelog.csv file exported from Redmine using ANSI encoding
and prepares the data for filling out the Excel timesheet template.

The script will:
1. Read the CSV file with ANSI (cp1252) encoding
2. Parse and clean the data
3. Prepare data structure for Excel template population
4. Display parsed data for verification

Requirements:
- pandas
- openpyxl (for future Excel operations)
"""

import os
import pandas as pd
from datetime import datetime
import logging

# Setup logging
logging.basicConfig(level=logging.INFO, format='%(asctime)s - %(levelname)s - %(message)s')
logger = logging.getLogger(__name__)


class RedmineCSVParser:
    def __init__(self, csv_filename='timelog.csv', verbose=False):
        """Initialize the CSV parser."""
        self.current_dir = os.path.dirname(os.path.abspath(__file__))
        self.csv_path = os.path.join(self.current_dir, csv_filename)
        self.excel_template_path = os.path.join(self.current_dir, 'Havi_elszamolas_2024_10 konyvelo.xlsx')
        self.parsed_data = None
        self.verbose = verbose
        
    def check_files_exist(self):
        """Check if required files exist."""
        if not os.path.exists(self.csv_path):
            raise FileNotFoundError(f"CSV file not found: {self.csv_path}")
        
        if not os.path.exists(self.excel_template_path):
            raise FileNotFoundError(f"Excel template not found: {self.excel_template_path}")
        
        if self.verbose:
            logger.info(f"Found CSV file: {self.csv_path}")
            logger.info(f"Found Excel template: {self.excel_template_path}")
    
    def read_csv_with_ansi(self):
        """Read the CSV file using ANSI (cp1252) encoding."""
        try:
            if self.verbose:
                logger.info("Reading CSV file with ANSI encoding...")
            
            # Try different encodings in order of preference
            encodings_to_try = ['cp1252', 'windows-1252', 'iso-8859-1', 'utf-8']
            
            for encoding in encodings_to_try:
                try:
                    df = pd.read_csv(self.csv_path, encoding=encoding)
                    if self.verbose:
                        logger.info(f"Successfully read CSV with {encoding} encoding")
                    return df
                except UnicodeDecodeError:
                    if self.verbose:
                        logger.warning(f"Failed to read with {encoding} encoding, trying next...")
                    continue
            
            # If all encodings fail, raise an error
            raise UnicodeDecodeError("Could not read CSV file with any of the attempted encodings")
            
        except Exception as e:
            logger.error(f"Error reading CSV file: {e}")
            raise
    
    def analyze_csv_structure(self, df):
        """Analyze and display the structure of the CSV data."""
        if self.verbose:
            logger.info("Analyzing CSV structure...")
            
            print("\n" + "="*60)
            print("CSV FILE ANALYSIS")
            print("="*60)
            
            print(f"Total rows: {len(df)}")
            print(f"Total columns: {len(df.columns)}")
            
            print("\nColumn names:")
            for i, col in enumerate(df.columns, 1):
                print(f"  {i:2d}. {col}")
            
            print("\nFirst 5 rows:")
            print(df.head().to_string())
            
            print("\nData types:")
            print(df.dtypes.to_string())
            
            print("\nSample of data (first row):")
            if len(df) > 0:
                for col in df.columns:
                    print(f"  {col}: {df.iloc[0][col]}")
        
        # Hardcoded Hungarian column names from Redmine CSV export
        expected_columns = {
            'Projekt': 'Project',
            'Dátum': 'Date', 
            'Felhasználó': 'User',
            'Aktivitás': 'Activity',
            'Feladat': 'Issue/Task',
            'Megjegyzés': 'Comment',
            'Óra': 'Hours'
        }
        
        if self.verbose:
            print("\nVerifying expected CSV columns:")
        found_columns = []
        missing_columns = []
        
        for hungarian_col, english_name in expected_columns.items():
            if hungarian_col in df.columns:
                found_columns.append(hungarian_col)
                if self.verbose:
                    print(f"  ✓ Found '{hungarian_col}' ({english_name})")
            else:
                missing_columns.append(hungarian_col)
                if self.verbose:
                    print(f"  ✗ Missing '{hungarian_col}' ({english_name})")
        
        if missing_columns:
            print(f"Warning: Missing expected columns: {missing_columns}")
        
        return found_columns
    
    def parse_time_entries(self, df):
        """Parse and structure the time entry data."""
        if self.verbose:
            logger.info("Parsing time entry data...")
        
        # This will be customized based on the actual CSV structure
        # For now, we'll create a generic parser that adapts to the columns found
        
        parsed_entries = []
        
        for index, row in df.iterrows():
            # Direct mapping using hardcoded Hungarian column names
            entry = {
                'project': row.get('Projekt', ''),
                'date': row.get('Dátum', ''),
                'user': row.get('Felhasználó', ''),
                'activity': row.get('Aktivitás', ''),
                'issue': row.get('Feladat', ''),
                'comment': row.get('Megjegyzés', ''),
                'hours': row.get('Óra', 0.0),
                # Keep original data for reference
                'original_data': {
                    'Projekt': row.get('Projekt', ''),
                    'Dátum': row.get('Dátum', ''),
                    'Felhasználó': row.get('Felhasználó', ''),
                    'Aktivitás': row.get('Aktivitás', ''),
                    'Feladat': row.get('Feladat', ''),
                    'Megjegyzés': row.get('Megjegyzés', ''),
                    'Óra': row.get('Óra', 0.0)
                }
            }
            
            parsed_entries.append(entry)
        
        return parsed_entries
    
    def summarize_data(self, parsed_entries):
        """Create summary statistics of the parsed data."""
        if not parsed_entries:
            if self.verbose:
                print("No data to summarize.")
            return
        
        if self.verbose:
            logger.info("Creating data summary...")
            print("\n" + "="*60)
            print("DATA SUMMARY")
            print("="*60)
        else:
            # Essential summary only
            total_hours = sum(float(str(entry['hours']).replace(',', '.')) for entry in parsed_entries if entry.get('hours'))
            print(f"📊 Processed {len(parsed_entries)} entries, {total_hours:.1f} total hours")
            return
        
        # Calculate total hours
        total_hours = 0
        valid_hours = 0
        
        for entry in parsed_entries:
            if 'hours' in entry and entry['hours']:
                try:
                    hours = float(str(entry['hours']).replace(',', '.'))
                    total_hours += hours
                    valid_hours += 1
                except (ValueError, TypeError):
                    pass
        
        print(f"Total entries: {len(parsed_entries)}")
        print(f"Entries with valid hours: {valid_hours}")
        print(f"Total hours: {total_hours:.2f}")
        
        # Show unique values for key fields
        for field in ['user', 'project', 'activity']:
            if field in parsed_entries[0]:
                unique_values = set()
                for entry in parsed_entries:
                    if entry.get(field):
                        unique_values.add(str(entry[field]))
                
                print(f"\nUnique {field}s:")
                for value in sorted(unique_values):
                    print(f"  - {value}")
        
        # Show date range
        dates = []
        for entry in parsed_entries:
            if 'date' in entry and entry['date']:
                dates.append(str(entry['date']))
        
        if dates:
            print(f"\nDate range:")
            print(f"  First entry: {min(dates)}")
            print(f"  Last entry: {max(dates)}")
    
    def fill_excel_template(self, parsed_entries):
        """Fill the Excel template with parsed CSV data."""
        from openpyxl import load_workbook
        from datetime import datetime, timedelta
        import calendar
        import re
        
        if self.verbose:
            logger.info("Loading Excel template...")
        
        try:
            # Load the Excel template
            wb = load_workbook(self.excel_template_path)
            ws = wb.active  # Get the active worksheet
            
            # Get user name (we know there's only one user)
            user_name = parsed_entries[0]['user'] if parsed_entries else "Unknown User"
            
            # Calculate first day of last month
            today = datetime.now()
            # Go to first day of current month, then subtract one day to get last month
            first_day_current_month = today.replace(day=1)
            last_month_last_day = first_day_current_month - timedelta(days=1)
            first_day_last_month = last_month_last_day.replace(day=1)
            
            if self.verbose:
                logger.info(f"User name: {user_name}")
                logger.info(f"Setting date range for: {first_day_last_month.strftime('%Y-%m')}")
            
            # Set A1 to first day of last month as a proper Excel date (date only, no time)
            # Store the original number format to preserve it
            original_a1_number_format = ws['A1'].number_format
            ws['A1'] = first_day_last_month.date()  # Set as date object, not datetime with time
            # Restore the original number format to preserve formatting
            ws['A1'].number_format = original_a1_number_format
            
            # Set B1 to user name
            ws['B1'] = user_name
            
            if self.verbose:
                logger.info(f"Set A1 to: {ws['A1'].value} (preserved original format: {original_a1_number_format})")
                logger.info(f"Set B1 to: {user_name}")
            
            # Create date lookup dictionary from CSV data
            date_lookup = {}
            for entry in parsed_entries:
                csv_date_str = entry['date']  # Format: "2025.09.29."
                try:
                    # Parse the CSV date format (2025.09.29.)
                    csv_date = datetime.strptime(csv_date_str.rstrip('.'), '%Y.%m.%d')
                    date_key = csv_date.strftime('%Y.%m.%d')  # Normalize format
                    
                    if date_key not in date_lookup:
                        date_lookup[date_key] = []
                    
                    # Extract task ID from issue field (find number after #)
                    task_id = ""
                    if entry['issue']:
                        match = re.search(r'#(\d+)', entry['issue'])
                        if match:
                            task_id = match.group(1)
                    
                    date_lookup[date_key].append({
                        'project': entry['project'],
                        'task_id': task_id,
                        'hours': entry['hours']
                    })
                    
                except ValueError as e:
                    if self.verbose:
                        logger.warning(f"Could not parse date '{csv_date_str}': {e}")
            
            if self.verbose:
                logger.info(f"Created date lookup for {len(date_lookup)} dates")
                
                # Check for dates with multiple entries and log them
                multiple_entry_dates = 0
                for date_key, entries in date_lookup.items():
                    if len(entries) > 1:
                        multiple_entry_dates += 1
                        total_hours_for_date = sum(entry['hours'] for entry in entries)
                        logger.info(f"Date {date_key} has {len(entries)} entries, summing to {total_hours_for_date}h")
                
                if multiple_entry_dates == 0:
                    logger.info("No dates with multiple entries found - each date has single entry")
                else:
                    logger.info(f"Found {multiple_entry_dates} dates with multiple entries - hours will be summed")
                
                logger.info("Clearing existing data in B3:F33...")
            for row in range(3, 34):  # Rows 3-33
                for col in ['B', 'C', 'D', 'E', 'F']:
                    ws[f'{col}{row}'] = None
            
            # Check and preserve formulas in column A (A3-A33)
            if self.verbose:
                logger.info("Checking column A for formulas...")
                for row in range(3, 6):  # Check first few rows as example
                    cell = ws[f'A{row}']
                    if hasattr(cell, 'data_type') and cell.data_type == 'f':
                        logger.info(f"A{row} contains formula: {cell.value}")
                    elif cell.value:
                        logger.info(f"A{row} contains value: {cell.value} (type: {type(cell.value)})")
                    else:
                        logger.info(f"A{row} is empty")
            
            # Process rows A3-A33 (31 days maximum)
            # Calculate all dates for the month based on A1
            days_in_month = calendar.monthrange(first_day_last_month.year, first_day_last_month.month)[1]
            
            for day in range(1, days_in_month + 1):
                row = day + 2  # A3 starts with day 1, A4 with day 2, etc.
                if row > 33:  # Don't exceed row 33
                    break
                    
                # Calculate the current date
                current_date = first_day_last_month.replace(day=day)
                date_key = current_date.strftime('%Y.%m.%d')
                
                # Check if it's a weekend
                is_weekend = current_date.weekday() >= 5  # Saturday=5, Sunday=6
                
                if date_key in date_lookup and not is_weekend:
                    # Fill work day data
                    entries_for_date = date_lookup[date_key]
                    
                    # If multiple entries for same date, sum the hours and combine info
                    total_hours = sum(entry['hours'] for entry in entries_for_date)
                    project = entries_for_date[0]['project']  # Use first project
                    task_ids = [entry['task_id'] for entry in entries_for_date if entry['task_id']]
                    task_id = task_ids[0] if task_ids else ""  # Use first task ID
                    
                    ws[f'B{row}'] = project           # Project name
                    # Convert task_id to integer to avoid "number stored as text" warning
                    if task_id:
                        try:
                            ws[f'C{row}'] = int(task_id)  # Task ID as number
                        except ValueError:
                            ws[f'C{row}'] = task_id       # Fallback to text if conversion fails
                    else:
                        ws[f'C{row}'] = None           # Empty if no task ID
                    ws[f'D{row}'] = "Fejlesztés"     # Activity
                    ws[f'E{row}'] = total_hours      # Hours
                    
                    if self.verbose:
                        logger.info(f"Row {row}: {date_key} -> {project}, #{task_id}, {total_hours}h")
                    
                elif is_weekend:
                    # Fill weekend data
                    ws[f'D{row}'] = "Munkaszüneti nap"
                    if self.verbose:
                        logger.info(f"Row {row}: {date_key} -> Weekend")
                
                # If no data for the date and not weekend, leave empty
            
            # Generate output filename: Havi_elszamolas_<year>_<month>_<User without spaces>.xlsx
            # Clean user name: remove spaces and replace special characters
            import unicodedata
            
            # Remove spaces and normalize unicode characters
            user_name_clean = user_name.replace(' ', '')
            
            # Replace common Hungarian accented characters
            char_replacements = {
                'á': 'a', 'à': 'a', 'â': 'a', 'ä': 'a', 'ã': 'a',
                'é': 'e', 'è': 'e', 'ê': 'e', 'ë': 'e',
                'í': 'i', 'ì': 'i', 'î': 'i', 'ï': 'i',
                'ó': 'o', 'ò': 'o', 'ô': 'o', 'ö': 'o', 'õ': 'o',
                'ú': 'u', 'ù': 'u', 'û': 'u', 'ü': 'u',
                'ő': 'o', 'ű': 'u',
                'Á': 'A', 'À': 'A', 'Â': 'A', 'Ä': 'A', 'Ã': 'A',
                'É': 'E', 'È': 'E', 'Ê': 'E', 'Ë': 'E',
                'Í': 'I', 'Ì': 'I', 'Î': 'I', 'Ï': 'I',
                'Ó': 'O', 'Ò': 'O', 'Ô': 'O', 'Ö': 'O', 'Õ': 'O',
                'Ú': 'U', 'Ù': 'U', 'Û': 'U', 'Ü': 'U',
                'Ő': 'O', 'Ű': 'U'
            }
            
            for accented, replacement in char_replacements.items():
                user_name_clean = user_name_clean.replace(accented, replacement)
            
            # Remove any remaining non-ASCII characters and special characters
            user_name_clean = ''.join(c for c in user_name_clean if c.isalnum())
            
            year = first_day_last_month.year
            month = first_day_last_month.month
            output_filename = f"Havi_elszamolas_{year}_{month:02d}_{user_name_clean}.xlsx"
            output_path = os.path.join(self.current_dir, output_filename)
            
            # Save the filled Excel file
            wb.save(output_path)
            
            if self.verbose:
                logger.info(f"Excel file saved as: {output_path}")
                print(f"\n✅ Excel template filled successfully!")
                print(f"📁 Saved as: {output_filename}")
                print(f"📅 Date range: {first_day_last_month.strftime('%Y-%m')}")
                print(f"👤 User: {user_name}")
                print(f"📊 Processed {len(date_lookup)} work days")
            else:
                print(f"✅ Excel saved: {output_filename}")
            
            return output_path
            
        except Exception as e:
            logger.error(f"Error filling Excel template: {e}")
            raise
    
    def run_analysis(self):
        """Run the complete CSV analysis process."""
        try:
            if self.verbose:
                logger.info("Starting CSV analysis...")
            
            # Check if files exist
            self.check_files_exist()
            
            # Read CSV file
            df = self.read_csv_with_ansi()
            
            # Analyze structure
            found_columns = self.analyze_csv_structure(df)
            
            # Parse the data
            parsed_entries = self.parse_time_entries(df)
            self.parsed_data = parsed_entries
            
            # Create summary
            self.summarize_data(parsed_entries)
            
            # Fill Excel template with parsed data
            output_path = self.fill_excel_template(parsed_entries)
            
            if self.verbose:
                print("\n" + "="*60)
                print("EXCEL TEMPLATE COMPLETED")
                print("="*60)
                print("The Excel timesheet has been successfully filled with CSV data.")
                print(f"Output file: {os.path.basename(output_path)}")
            
            return parsed_entries
            
        except Exception as e:
            logger.error(f"Analysis failed: {e}")
            raise


def main():
    """Main function to run the CSV analysis."""
    import sys
    
    # Check for verbose flag
    verbose = '--verbose' in sys.argv or '-v' in sys.argv
    
    try:
        parser = RedmineCSVParser(verbose=verbose)
        parsed_data = parser.run_analysis()
        
        if not verbose:
            print("✅ Timesheet processing completed!")
        else:
            print(f"\nSuccessfully parsed {len(parsed_data)} entries from the CSV file.")
        
    except Exception as e:
        print(f"Error: {e}")
        return 1
    
    return 0


if __name__ == '__main__':
    exit(main())